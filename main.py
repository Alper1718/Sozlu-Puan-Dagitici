import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
from openpyxl import load_workbook
import random

FILE_NAME = "PERFORMANS ÖLÇEK sınıf listeli olan SON.xlsx"
CRITERIA_MAX_VALUES = [40, 10, 20, 20, 10]

workbook = load_workbook(FILE_NAME)
sheet = workbook.active

# Scrolling wheels are handled differently on different operating systems.
def _on_mousewheel(event):
    if event.num == 4:  # I use Arch by the way.
        canvas.yview_scroll(-1, "units")
    elif event.num == 5:  # Linux scroll down
        canvas.yview_scroll(1, "units")
    else:  # Windows and macOS
        canvas.yview_scroll(-1 * (event.delta // 120), "units")


def resize_canvas(event=None):
    canvas_width = canvas.winfo_width()
    scrollable_frame_width = scrollable_frame.winfo_reqwidth()
    canvas.itemconfigure(
        canvas.create_window(
            (canvas_width // 2 - scrollable_frame_width // 2, 0),
            window=scrollable_frame,
            anchor="n",
        )
    )

def distribute_points(final_points):
    points = [0] * len(CRITERIA_MAX_VALUES)
    if final_points % 5 != 0:
        raise ValueError("Final points must be a multiple of 5.")
    remaining_points = final_points
    while remaining_points > 0:
        idx = random.randint(0, len(CRITERIA_MAX_VALUES) - 1)
        if points[idx] + 5 <= CRITERIA_MAX_VALUES[idx] and remaining_points >= 5:
            points[idx] += 5
            remaining_points -= 5
    return points

def distribute_exact_points():
    return CRITERIA_MAX_VALUES.copy()

def save_to_excel(name, final_points, points):
    row_idx = names.index(name) + 8
    for col, pts in enumerate(points, start=4):
        sheet.cell(row=row_idx, column=col, value=pts)
    workbook.save(FILE_NAME)

def submit():
    try:
        for name, entry in zip(names, note_entries):
            note_text = entry.get()
            if not note_text:
                continue
            final_points = int(note_text)
            if final_points < 0 or final_points > 100:
                raise ValueError(f"Final points for {name} must be between 0 and 100.")
            points = distribute_exact_points() if final_points == 100 else distribute_points(final_points)
            save_to_excel(name, final_points, points)
        messagebox.showinfo("Success", "All data saved successfully!")
        root.destroy()
    except ValueError as e:
        messagebox.showerror("Invalid Input", str(e))

names = [
    f"{sheet.cell(row=row, column=2).value} {row-7}-) {sheet.cell(row=row, column=3).value}"
    for row in range(8, sheet.max_row + 1)
    if sheet.cell(row=row, column=3).value
]

root = tk.Tk()
root.title("Performance Notes")
root.state("normal")
root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}")
root.configure(bg=None)  # Remove background color to use default theme

title_label = tk.Label(root, text="Performance Notes", font=("Arial", 24, "bold"))
title_label.pack(side="top", pady=root.winfo_screenheight() * 0.02)

canvas_frame = ttk.Frame(root)
canvas_frame.pack(side="top", fill="both", expand=True)

canvas = tk.Canvas(canvas_frame)
scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
scrollable_frame = ttk.Frame(canvas)

canvas.create_window((0, 0), window=scrollable_frame, anchor="n")

canvas.configure(yscrollcommand=scrollbar.set)
canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

canvas.bind("<Configure>", resize_canvas)
scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

canvas.bind_all("<MouseWheel>", _on_mousewheel)  # Windows and macOS
canvas.bind_all("<Button-4>", _on_mousewheel)  # Linux scroll up
canvas.bind_all("<Button-5>", _on_mousewheel)  # Linux scroll down

note_entries = []

for i, name in enumerate(names):
    name_label = ttk.Label(scrollable_frame, text=name, font=("Arial", 12))
    name_label.grid(row=i, column=0, sticky=tk.W, pady=5, padx=10)

    note_entry = ttk.Entry(scrollable_frame, width=10)
    note_entry.grid(row=i, column=1, pady=5, padx=10)
    note_entries.append(note_entry)

submit_image = Image.open("submit.png").resize((40, 40))
submit_photo = ImageTk.PhotoImage(submit_image)
submit_button = ttk.Button(root, text=" Submit", image=submit_photo, compound="left", command=submit)
submit_button.pack(side="bottom", pady=root.winfo_screenheight() * 0.02)

left_image = Image.open("left_image.png").resize((200, 200))
left_photo = ImageTk.PhotoImage(left_image)
left_label = tk.Label(root, image=left_photo)
left_label.place(relx=0.05, rely=0.95, anchor="sw")

right_image = Image.open("right_image.png").resize((200, 200))
right_photo = ImageTk.PhotoImage(right_image)
right_label = tk.Label(root, image=right_photo)
right_label.place(relx=0.95, rely=0.95, anchor="se")

scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

root.mainloop()
